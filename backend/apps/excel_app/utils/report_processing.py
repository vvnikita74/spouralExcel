import os
import inspect
import re
import pymorphy2
import openpyxl
from django.conf import settings
from datetime import datetime
from apps.excel_app.models import Sheet


# Python 3.13 bug fix
def patched_getargspec(func):
    fullargspec = inspect.getfullargspec(func)
    return fullargspec.args, fullargspec.varargs, fullargspec.varkw, fullargspec.defaults


inspect.getargspec = patched_getargspec

morph = pymorphy2.MorphAnalyzer()


def get_gender(word):
    parse = morph.parse(word)[0]

    if 'masc' in parse.tag:
        return 'masc'
    elif 'femn' in parse.tag:
        return 'femn'
    elif 'neut' in parse.tag:
        return 'neut'

    return None


def change_gender(word, target_gender):
    parse = morph.parse(word)[0]
    inflected = parse.inflect({target_gender})
    return inflected.word if inflected else None


def generate_report(data, username):
    template_path = os.path.join(os.path.dirname(__file__), 'report.xlsx')
    wb = openpyxl.load_workbook(template_path)
    count = 0
    for sheet in Sheet.objects.all():
        ws = wb.worksheets[sheet.index]
        count += 1
        for cell_data in sheet.get_data():
            cell = ws[cell_data.index]
            template = cell_data.template
            template = substitute_placeholders(template, data)
            cell.value = template

        if sheet.countCell:
            ws[sheet.countCell] = count

        sheet.save()

    if wb.worksheets:
        wb.remove(wb.worksheets[-1])

    report_dir = os.path.join(settings.MEDIA_ROOT, 'reports')
    os.makedirs(report_dir, exist_ok=True)
    report_full_filename = os.path.join(report_dir,
                                        f'{datetime.now().strftime("%H-%M_%d.%m.%Y")}-{username}.xlsx')
    wb.save(report_full_filename)
    report_filename = os.path.splitext(os.path.basename(report_full_filename))[0]
    # os.system(f'libreoffice --headless --convert-to pdf --outdir'
    #           f' {report_dir} {report_filename}')
    return report_filename


def substitute_placeholders(template, data):
    gentArr = {
        "Родительный": "gent",
        "Именительный": "nomn",
        'Дательный': 'datv',
        'Винительный': 'accs',
        'Творительный': 'ablt',
        'Предложный': 'loct',
        'Звательный': 'voct',
    }

    def replace_match(match):
        key = match.group(1)
        keyArr = key.split('.')
        length = len(keyArr)
        match length:
            case 1:
                return data.get(key, f'${key}$')
            case 2:
                word, key = keyArr
                if key in gentArr:
                    initial = data.get(word, '')
                    gender = get_gender(initial.split(' ')[-1])
                    if gender:
                        inflected_word = change_gender(word, gender)
                        if inflected_word:
                            return inflected_word
                elif key in data:
                    initial = data[key]
                    gender = get_gender(initial.split(' ')[-1])
                    if gender:
                        inflected_word = change_gender(word, gender)
                        if inflected_word:
                            return inflected_word
            case 3:
                # Обработка случая с тремя частями в ключе
                pass

        return data.get(key, f'${key}$')

    return re.sub(r'\$(\w+(\.\w+)*)\$', replace_match, template)
