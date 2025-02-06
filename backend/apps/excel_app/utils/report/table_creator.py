import json
import os
from enum import Enum
from openpyxl.styles import Border, Side, Alignment, Font
from typing import List, Any

from .data_models import Documentation, UniversalObject, Section, Defect


class TableType(Enum):
    DOCUMENTATION = 'documentation'
    CONTENT = 'content'
    DEFAULT = 'default'
    DEFECTS = 'defects'


class Table:
    def __init__(self, table_type: TableType):
        """
        Инициализирует объект Table с заданным типом таблицы.

        :param table_type: Тип таблицы (TableType)
        """
        self.table_type = table_type

    @staticmethod
    def set_border(ws, top_left: str, bottom_right: str):
        """
        Устанавливает границы для диапазона ячеек.

        :param ws: Рабочий лист Excel
        :param top_left: Левая верхняя ячейка диапазона
        :param bottom_right: Правая нижняя ячейка диапазона
        """
        thin_border = Border(
            left=Side(border_style='thin', color='000000'),
            right=Side(border_style='thin', color='000000'),
            top=Side(border_style='thin', color='000000'),
            bottom=Side(border_style='thin', color='000000')
        )

        min_col, min_row = ws[top_left].column, ws[top_left].row
        max_col, max_row = ws[bottom_right].column, ws[bottom_right].row

        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border

    @staticmethod
    def create_table(ws, cell_data) -> (int, int, int):
        """
        Создает таблицу на рабочем листе Excel.

        :param ws: Рабочий лист Excel
        :param cell_data: Объект данных ячейки
        :return: Начальные столбец и строка, вертикальный зазор
        """
        start_cell = cell_data.index
        start_col = ws[start_cell].column
        start_row = ws[start_cell].row
        vertical_gap = cell_data.verticalGap

        for cell_info in cell_data.cells:
            width = cell_info.get('width')
            end_col = start_col + width - 1
            end_row = start_row + vertical_gap - 1
            Table.apply_cell_styles(ws, start_row, start_col, end_row, end_col)
            try:
                value = cell_info.get('title')
            except KeyError:
                value = ''
            ws.cell(row=start_row, column=start_col, value=value)
            start_col = end_col + 1

        start_row += vertical_gap
        start_col = ws[start_cell].column
        return start_col, start_row, vertical_gap

    @staticmethod
    def apply_cell_styles(ws, start_row: int, start_col: int, end_row: int,
                          end_col: int):
        """
        Вычисляет границы, устанавливает границы, объединяет ячейки и применяет стили.

        :param ws: Рабочий лист Excel
        :param start_row: Начальная строка
        :param start_col: Начальный столбец
        :param end_row: Конечная строка
        :param end_col: Конечный столбец
        """
        Table.set_border(ws,
                         f'{ws.cell(row=start_row, column=start_col).coordinate}',
                         f'{ws.cell(row=end_row, column=end_col).coordinate}')
        ws.merge_cells(start_row=start_row, start_column=start_col,
                       end_row=end_row, end_column=end_col)

    def fill_table(self, ws, cell_data, input_value: str):
        """
        Обрабатывает данные и вставляет их в соответствующие ячейки.

        :param ws: Рабочий лист Excel
        :param cell_data: Объект данных ячейки
        :param input_value: JSON строка с данными/список объектов
        """
        data_objects = self._get_data_objects(input_value)
        max_sheet_id = get_max_sheet_id(input_value)
        if data_objects:
            start_col, start_row, vertical_gap = Table.create_table(ws,
                                                                    cell_data)
            for idx, obj in enumerate(data_objects):
                for cell_info in cell_data.cells:
                    key = cell_info.get('key')
                    width = cell_info.get('width')
                    end_col = start_col + width - 1
                    end_row = start_row + vertical_gap - 1
                    try:
                        value = idx + 1 if self.table_type == TableType.DOCUMENTATION and key == 'index' else getattr(
                            obj, key, '')
                    except AttributeError:
                        value = ''
                    ws.cell(row=start_row, column=start_col, value=value)
                    Table.apply_cell_styles(ws, start_row, start_col, end_row,
                                            end_col)
                    start_col = end_col + 1

                start_row += vertical_gap
                start_col = ws[cell_data.index].column
            if max_sheet_id and cell_data.listsCell:
                ws[cell_data.listsCell].value = max_sheet_id

    def fill_defects(self, ws, cell_data, data, sheet):
        # Извлечение данных из sheet.data
        sheet_data = sheet.get_data()
        code_cell = next(
            (item for item in sheet_data if item.inputKey == 'code'), None)
        report_date_cell = next(
            (item for item in sheet_data if item.inputKey == 'reportDate'),
            None)

        # Проверка наличия данных
        if not code_cell or not report_date_cell:
            raise ValueError("Не найдены необходимые данные в sheet.data")

        code_value = data.get(code_cell.inputKey, '')
        report_date_value = data.get(report_date_cell.inputKey, '')

        defects = self._create_defects(cell_data, data)
        start_cell = cell_data.index
        start_col = ws[start_cell].column
        start_row = ws[start_cell].row
        sheet_index = 1
        last_inserted_sheet = ws
        original_sheet = ws.parent.copy_worksheet(
            ws)  # Сохранение ссылки на оригинальный лист
        added_defects = set()  # Отслеживание добавленных дефектов

        for defect in defects:
            if defect in added_defects:
                continue  # Пропуск дефекта, если он уже был добавлен

            if not self.check_table_height(ws, cell_data, start_row, start_col,
                                           defect):
                # Создание копии оригинального листа
                new_ws = original_sheet
                new_ws.title = f"{original_sheet.title.replace('Copy', '')}{sheet_index}"  # Название нового листа с индексом

                # Перемещение нового листа сразу после последнего вставленного листа
                original_index = ws.parent.index(last_inserted_sheet)
                ws.parent._sheets.remove(new_ws)
                ws.parent._sheets.insert(original_index + 1, new_ws)

                ws = new_ws
                last_inserted_sheet = new_ws  # Обновление последнего вставленного листа
                start_row = ws[
                    start_cell].row  # Сброс start_row для нового листа
                sheet_index += 1

                if sheet.countCell:
                    ws[sheet.countCell] = sheet.index + sheet_index

                # Вставка значений в соответствующие ячейки
                ws[code_cell.index] = code_value
                ws[report_date_cell.index] = report_date_value

            header_name = cell_data.cells.get('names').get(defect.type)
            start_row = self.draw_table_header(ws, cell_data, start_row,
                                               start_col, header_name)
            start_row = self.draw_table_elements(ws, cell_data, start_row,
                                                 start_col, defect)
            added_defects.add(defect)  # Пометка дефекта как добавленного

    @staticmethod
    def check_table_height(ws, cell_data, start_row, start_col, defect):
        total_height = 0
        defect_dict = defect.to_dict()
        for row in cell_data.cells['values']['rows']:
            if isinstance(defect_dict.get(row['key'], ''), list):
                value_height = cell_data.cells['header']['minHeight'] + len(
                    defect_dict.get(row['key'], '')) - 1
            else:
                value_height = cell_data.cells['values']['valueHeight']
            total_height += value_height

        end_row = start_row + total_height - 1
        max_cell = cell_data.cells["maxCell"]
        max_row = ws[max_cell].row
        return end_row <= max_row

    @staticmethod
    def draw_table_header(ws, cell_data, start_row, start_col, header_name):
        header_length = 35
        end_col = start_col + header_length
        end_row = start_row + 1

        # Вставка значения для заголовка
        cell = ws.cell(row=start_row, column=start_col + 1,
                       value=header_name)  # Сдвиг на одну ячейку вправо

        # Центрирование значения, установка жирного шрифта и размера шрифта 14
        cell.alignment = Alignment(horizontal='center', vertical='center',
                                   wrapText=True)
        cell.font = Font(name="Arial", bold=True, size=12)

        # Объединение ячеек для заголовка
        ws.merge_cells(start_row=start_row, start_column=start_col + 1,
                       end_row=end_row, end_column=end_col)

        return end_row + 1  # Возвращаем следующую строку после заголовка

    @staticmethod
    def draw_table_elements(ws, cell_data, start_row, start_col, defect):
        defect_dict = defect.to_dict()
        for row in cell_data.cells['values']['rows']:
            # Вычисление конечной строки для ячейки имени
            name_end_row = start_row + cell_data.cells['header'][
                'nameHeight'] - 1
            name_end_col = start_col + cell_data.cells['header']['width'] - 1

            # Вставка значения для ячейки имени
            name_cell = ws.cell(row=start_row, column=start_col,
                                value=row['name'])
            name_cell.alignment = Alignment(vertical="top",
                                            horizontal="right", wrap_text=True)

            # Вычисление начального столбца для ячейки значения
            value_start_col = start_col + cell_data.cells['header']['width']
            # Получение значения для дефекта
            value = defect_dict.get(row['key'], '')
            if isinstance(value, list):
                value_height = cell_data.cells['header']['minHeight'] + len(
                    value) - 1
                value = '\n'.join(
                    f"- {item}" if i == 0 else f"- {item}" for i, item in
                    enumerate(filter(None,
                                     value)))  # Преобразование списка в маркеры
            else:
                value_height = cell_data.cells['values']['valueHeight']

            # Вычисление конечной строки для ячейки значения
            value_end_row = start_row + value_height - 1
            value_end_col = value_start_col + cell_data.cells['values'][
                'width'] - 1

            # Установка границ для ячейки имени
            Table.set_border(ws, name_cell.coordinate,
                             ws.cell(row=value_end_row,
                                     column=name_end_col).coordinate)

            # Объединение ячеек для имени
            ws.merge_cells(start_row=start_row, start_column=start_col,
                           end_row=value_end_row, end_column=name_end_col)

            # Вставка значения для ячейки значения
            value_cell = ws.cell(row=start_row, column=value_start_col,
                                 value=value)
            value_cell.alignment = Alignment(wrap_text=True, horizontal="left",
                                             vertical="top")

            # Установка границ для ячейки значения
            Table.set_border(ws, value_cell.coordinate,
                             ws.cell(row=value_end_row,
                                     column=value_end_col).coordinate)

            # Объединение ячеек для значения
            ws.merge_cells(start_row=start_row, start_column=value_start_col,
                           end_row=value_end_row, end_column=value_end_col)

            # Переход к следующей строке после текущей строки
            start_row = max(name_end_row, value_end_row) + 1

        return start_row + 1  # Возвращаем следующую строку после элементов таблицы

    def _create_defects(self, cell_data, data):
        defects = []
        for key in cell_data.cells['names'].keys():
            defect_data = json.loads(data.get(key, '{}'))
            if defect_data:  # Check if defect_data is not empty
                defect = Defect.from_dict(type=key, data=defect_data)
                defects.append(defect)
        return defects

    def _get_data_objects(self, input_value: Any) -> List[Any] | None:
        """
        Возвращает список объектов данных в зависимости от типа таблицы.

        :param input_value: JSON строка с данными/список объектов
        :return: Список объектов данных
        """
        if self.table_type == TableType.DOCUMENTATION:
            if input_value:
                return [Documentation.from_dict(doc) for doc in
                        json.loads(input_value)]
            else:
                return None
        elif self.table_type == TableType.CONTENT:
            return input_value
        else:
            return [UniversalObject.from_dict(obj) for obj in
                    json.loads(input_value)]


def get_max_sheet_id(sections):
    """
    Возвращает максимальный sheetId среди списка секций.

    :param sections: Список объектов Section
    :return: Максимальный sheetId
    """
    if not isinstance(sections, list) or not all(
            isinstance(section, Section) for section in sections):
        return None
    if not sections:
        return None
    return max(section.sheetId for section in sections)
