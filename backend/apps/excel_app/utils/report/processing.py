import copy
import json
import os
import re
from typing import List
import pymorphy2
import openpyxl
from django.conf import settings

from apps.excel_app.models import Sheet
from apps.excel_app.utils.report.table_creator import Table, TableType
from apps.excel_app.utils.report.data_models import Section
from apps.excel_app.utils.report.constant_tags import gender_tags, \
    month_tags, case_tags, register_tags

from apps.excel_app.utils.morph_patch import apply_patch

from apps.excel_app.utils.report.media_processing import (insert_media,
                                                          MediaParams)

from apps.excel_app.utils.report.media_processing import process_images

apply_patch()
morph = pymorphy2.MorphAnalyzer()


def generate_report(data, filename, user_files):
    """
    Генерирует отчет на основе данных и сохраняет его в файл.

    :param data: Входные данные
    :return: Имя файла отчета
    """
    template_path = os.path.join(os.path.dirname(__file__), '..', '..', '..',
                                 '..', 'static', 'report.xlsx')
    template_path = os.path.abspath(template_path)
    wb = openpyxl.load_workbook(template_path)

    count = 0
    sections = []
    sheets = sorted(Sheet.objects.all(), key=lambda sheet: sheet.index)
    sheets_copy = copy.deepcopy(sheets)
    content_cell_data = None
    inserted_sheets_count = 0
    indices_updated = False
    for sheet in sheets_copy:
        print(f"Processing sheet {sheet.name} with index "
              f"{count} + {inserted_sheets_count}")

        ws = wb.worksheets[sheet.index]
        count += 1
        if sheet.countCell:
            print(f"Setting count cell to {count + inserted_sheets_count} to list {sheet.name}")
            ws[sheet.countCell] = count + inserted_sheets_count

        for cell_data in sheet.get_data():
            content_cell_data, inserted_sheets_count = process_cell_data(ws,
                                                                         cell_data,
                                                                         data,
                                                                         content_cell_data,
                                                                         sheet,
                                                                         inserted_sheets_count,
                                                                         user_files)
        if sheet.contentSection:
            sections.extend(process_sections(sheet, count))

            # Обновление индексов всех последующих листов only once
        if inserted_sheets_count > 0 and not indices_updated:
            for s in sheets_copy:
                if s.index > sheet.index:
                    s.index += inserted_sheets_count
            indices_updated = True
    if wb.worksheets:
        wb.remove(wb.worksheets[-1])

    if content_cell_data:
        content_ws = wb['Содержание']
        table = Table(TableType.CONTENT)
        table.fill_table(content_ws, content_cell_data, sections)

    else:
        print("Warning: content_cell_data is None, skipping fill_content")

    # print("Sheets in workbook after processing:")
    # for index, sheet in enumerate(wb.sheetnames):
    #     print(f"Sheet name: {sheet}, Index: {index}")
    save_report(wb, filename)


def process_cell_data(ws, cell_data, data, content_cell_data, sheet,
                      inserted_sheets_count, user_files):
    cell = ws[cell_data.index]
    template = cell_data.template
    input_value = get_nested_value(data, cell_data.inputKey,
                                   cell_data.defaultValue)
    if cell_data.tableType:
        match cell_data.tableType:
            case 'documentation':
                table = Table(TableType(cell_data.tableType))
                table.fill_table(ws, cell_data, input_value)
            case "defects":
                table = Table(TableType(cell_data.tableType))
                inserted_sheets_count = table.fill_defects(ws, cell_data,
                                                           data, sheet,
                                                           inserted_sheets_count)
            case 'content':
                content_cell_data = cell_data
            case 'media':
                media_params = MediaParams(cell_data, data, sheet)
                insert_media(ws, media_params)
            case "images":
                inserted_sheets_count = process_images(ws, cell_data, data,
                                                       sheet,
                                                       inserted_sheets_count,
                                                       user_files)
    else:
        try:
            if not template:
                cell.value = cell_data.defaultValue if not input_value else input_value
            else:
                template = substitute_placeholders(template, data)
                cell.value = template
        except AttributeError:
            cell.value = ""

    return content_cell_data, inserted_sheets_count


def save_report(wb: openpyxl.Workbook, filename: str) -> None:
    """
    Сохраняет отчет в файл и конвертирует его в PDF.

    :param wb: Рабочая книга Excel
    :param filename: Имя файла отчета
    """
    report_dir = os.path.join(settings.MEDIA_ROOT, 'reports')
    os.makedirs(report_dir, exist_ok=True)
    os_filename = os.path.join(report_dir, filename + '.xlsx')

    try:
        wb.save(os_filename)
        # print(f'Отчет сохранен в {os_filename}')
    except Exception as e:
        print(f'Ошибка при сохранении отчета: {e}')
    os.system(
        f'libreoffice --headless --convert-to pdf --outdir {report_dir} {os_filename}')


def process_sections(sheet: Sheet, count: int) -> List[Section]:
    """
    Обрабатывает секции листа.

    :param sheet: Лист Excel
    :param count: Счетчик листов
    :return: Список секций
    """
    sections = [Section(sectionId=sheet.section, sectionName=sheet.name,
                        sheetId=count)]
    if sheet.subsections:
        subsections = json.loads(sheet.subsections)
        for subsection in subsections:
            sections.append(Section(sectionId=subsection.get('sectionId'),
                                    sectionName=subsection.get('sectionName'),
                                    sheetId=count))
    return sections


def get_nested_value(data, key, default_value=None):
    """
    Получает вложенное значение из словаря по ключу.

    :param data: Словарь данных
    :param key: Ключ для поиска
    :param default_value: Значение по умолчанию
    :return: Значение по ключу
    """
    keys = key.split('.')
    value = data
    for k in keys:
        if isinstance(value, str):
            try:
                value = json.loads(value)
            except json.JSONDecodeError:
                return default_value
        if not isinstance(value, dict):
            return default_value
        value = value.get(k, None)
        if value is None:
            return default_value
    return value


def substitute_placeholders(template, data):
    """
    Заменяет плейсхолдеры в шаблоне на значения из данных.

    :param template: Шаблон с плейсхолдерами
    :param data: Словарь данных
    :return: Шаблон с замененными плейсхолдерами
    """

    def replace_match(match):
        key = match.group(1)
        key_arr = key.split('.')
        length = len(key_arr)
        inflected_word = ''
        match length:
            case 2:
                word, key = key_arr

                if key in case_tags:
                    initial = data.get(word, '')
                    inflected_word = change_case(initial, key)
                elif key in data:
                    initial = data[key]
                    inflected_word = change_case(word, get_gender(
                        initial.split(' ')[-1]))
                elif key in register_tags:
                    initial = data.get(word, '')
                    match key:
                        case 'upper':
                            inflected_word = initial.upper()
                        case 'lower':
                            inflected_word = initial.lower()
                        case _:
                            inflected_word = initial
                elif word in data:
                    try:
                        word_data = json.loads(data[word])
                        inflected_word = word_data.get(key)
                    except Exception as e:
                        print(f"Error: {e}")
                if inflected_word:
                    return str(inflected_word)
            case 3:
                word, key1, key2 = key_arr
                if key1 in case_tags:
                    initial = data.get(word, '')
                    inflected_word = change_case(initial, key1)
                elif key1 in data:
                    initial = data[key1]
                    inflected_word = change_case(word, get_gender(
                        initial.split(' ')[-1]))
                if key2 in register_tags:
                    match key2:
                        case 'upper':
                            inflected_word = inflected_word.upper()
                        case 'lower':
                            inflected_word = inflected_word.lower()
                        case _:
                            inflected_word = inflected_word
                else:  # Случай, когда второй ключ не является тегом регистра
                    pass
                if inflected_word:
                    return str(inflected_word)
            case _:
                value = data.get(key, f'${key}$')
                if value.split('.')[0] in month_tags:
                    month = month_tags[value.split('.')[0]]
                    year = value.split('.')[1]
                    match len(year):
                        case 2:
                            return f'{month} 20{year} года'
                        case _:
                            return f'{month} {year} года'
                return str(value)

    return re.sub(r'\$(\w+(\.\w+)*)\$', replace_match, template)


def get_gender(word):
    """
    Определяет род слова.

    :param word: Слово для анализа
    :return: Род слова
    """
    parse = morph.parse(word)[0]

    for gender, tag in gender_tags.items():
        if gender in parse.tag:
            return tag

    return None


def change_case(phrase, target_case):
    """
    Изменяет род слова и падеж слова на заданные.

    :param phrase: Фраза для изменения
    :param target_case: Целевой падеж
    :return: Фраза с измененным падежом
    """
    words = phrase.split(' ')
    inflected_words = []
    for word in words:
        was_capitalized = word.istitle()  # Проверяем, была ли первая буква заглавной
        parsed_word = morph.parse(word)[0]
        inflected_word = parsed_word.inflect({target_case})
        if inflected_word:
            inflected_word_str = inflected_word.word
            if was_capitalized:
                inflected_word_str = inflected_word_str.capitalize()  # Восстанавливаем регистр
            inflected_words.append(inflected_word_str)

    if len(inflected_words) > 1:
        gender = get_gender(inflected_words[-1])
        parsed_word = morph.parse(inflected_words[0])[0]
        inflected_words[0] = parsed_word.inflect({gender}).word
    return ' '.join(inflected_words)
