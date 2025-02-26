import json
import os
from io import BytesIO

from openpyxl.styles import Alignment
from openpyxl.utils.cell import (coordinate_from_string,
                                 column_index_from_string, get_column_letter)

from PIL import Image as PILImage, ImageDraw, ImageFont
from django.conf import settings
from openpyxl.drawing.image import Image


def insert_media(ws, media_params):
    """
    Вставляет изображение в указанный диапазон ячеек с выравниванием по центру.

    :param ws: Рабочий лист Excel
    :param media_params: Параметры изображения
    """
    # Загружаем изображение
    img = PILImage.open(media_params.file)
    # Задаем размеры изображения
    img = img.resize((media_params.image_width, media_params.image_height),
                     PILImage.LANCZOS)
    draw = ImageDraw.Draw(img)
    # Коэффициент масштабирования
    scale_factor = media_params.print_scale
    font_path = os.path.join(settings.STATIC_ROOT, 'fonts', 'Arial.ttf')
    try:
        # Шрифты с учетом коэффициента масштабирования
        font_1 = ImageFont.truetype(font_path,
                                    int(media_params.text_font_size1 *
                                        scale_factor))
        font_2 = ImageFont.truetype(font_path,
                                    int(media_params.text_font_size2 * scale_factor))
    except Exception as e:
        print(f"Шрифт {font_path} не найден")
        return
    # Текст для печати
    text1 = media_params.textLine1
    text2 = media_params.textLine2
    text3 = media_params.textLine3
    text_color = media_params.text_color
    frame_color = media_params.frame_color

    # Размеры изображения и текста
    img_width, img_height = img.size
    text1_bbox = draw.textbbox((0, 0), text1, font=font_1)
    text2_bbox = draw.textbbox((0, 0), text2, font=font_2)
    text3_bbox = draw.textbbox((0, 0), text3, font=font_2)
    text1_width = text1_bbox[2] - text1_bbox[0]
    text1_height = text1_bbox[3] - text1_bbox[1]
    text2_width = text2_bbox[2] - text2_bbox[0]
    text2_height = text2_bbox[3] - text2_bbox[1]
    text3_width = text3_bbox[2] - text3_bbox[0]
    text3_height = text3_bbox[3] - text3_bbox[1]

    # Общая высота и ширина текста с учетом коэффициента масштабирования
    total_text_height = text1_height + text2_height + text3_height + int(
        media_params.line_spacing * 2 * scale_factor)  # line_spacing -
    # промежутки между строками

    # Позиция текста
    text_x = img_width * int(media_params.print_x) / 100
    text_y = img_height * int(media_params.print_y) / 100
    # Ширина рамки
    frame_margin = media_params.frame_margin
    max_text_width = max(text1_width, text2_width, text3_width)
    frame_x_len = max_text_width + 2 * frame_margin  # 2 * frame_margin -
    # отступы по бокам
    frame_width = media_params.frame_width

    # Центрирование текста относительно рамки
    text1_x = text_x + (frame_x_len - text1_width) / 2
    text2_x = text_x + (frame_x_len - text2_width) / 2
    text3_x = text_x + (frame_x_len - text3_width) / 2

    # Рисуем текст
    draw.text((text1_x, text_y), text1, fill=text_color, font=font_1)
    draw.text((text2_x, text_y + text1_height + int(
        media_params.line_spacing * scale_factor)), text2,
              fill=text_color, font=font_2)
    draw.text((text3_x,
               text_y + text1_height + text2_height + int(
                   media_params.line_spacing * 2 * scale_factor)),
              text3, fill=text_color, font=font_2)

    # Рисуем рамку
    draw.rectangle([(text_x - frame_margin, text_y - frame_margin),
                    (text_x + frame_x_len,
                     text_y + total_text_height + 2 * frame_margin)],
                   outline=frame_color, width=frame_width)

    # Сохраняем измененное изображение в BytesIO
    img_byte_arr = BytesIO()
    img.save(img_byte_arr, format='PNG')
    img_byte_arr.seek(0)
    try:
        # Вставляем изображение в начальную ячейку диапазона
        img = Image(img_byte_arr)
        ws.add_image(img, media_params.start_cell)
        ws.merge_cells(f"{media_params.start_cell}:{media_params.merge_cell}")
    except Exception as e:
        print(f"Ошибка при вставке изображения: {e}")


class MediaParams:
    def __init__(self, cell_data, data, sheet):
        self.file = sheet.files
        self.start_cell = cell_data.index
        self.merge_cell = cell_data.cells["mergeCell"]
        self.textLine3 = data.get(cell_data.cells["textLine3Key"])
        self.print_x = cell_data.cells["printX"]
        self.print_y = cell_data.cells["printY"]
        self.image_width = int(cell_data.cells["imageWidth"])
        self.image_height = int(cell_data.cells["imageHeight"])
        self.print_scale = int(cell_data.cells["printScale"])
        self.frame_margin = int(cell_data.cells["frameMargin"])
        self.frame_width = int(cell_data.cells["frameWidth"])
        self.line_spacing = int(cell_data.cells["lineSpacing"])
        self.text_color = eval(cell_data.cells["textColor"])
        self.frame_color = eval(cell_data.cells["frameColor"])
        self.textLine1 = cell_data.cells["textLine1"]
        self.textLine2 = cell_data.cells["textLine2"]
        self.text_font_size1 = int(cell_data.cells["textFontSize1"])
        self.text_font_size2 = int(cell_data.cells["textFontSize2"])

    def __str__(self):
        return (f"MediaParams(file={self.file}, start_cell={self.start_cell}, "
                f"merge_cell={self.merge_cell}, textLine3={self.textLine3}, "
                f"print_x={self.print_x}, print_y={self.print_y}, "
                f"image_width={self.image_width}, image_height={self.image_height}, "
                f"print_scale={self.print_scale}, frame_margin={self.frame_margin}, "
                f"frame_width={self.frame_width}, line_spacing={self.line_spacing}, "
                f"text_color={self.text_color}, frame_color={self.frame_color}, "
                f"textLine1={self.textLine1}, textLine2={self.textLine2}, "
                f"text_font_size1={self.text_font_size1}, "
                f"text_font_size2={self.text_font_size2})")


class UserImage:
    def __init__(self, image, description):
        self.image = image
        self.description = description

    def __str__(self):
        return f"UserImage(image={self.image}, description={self.description})"


def get_image_description(image_key, cell_data, data):
    """
    Получает описание изображения.

    :param image: Изображение
    :param cell_data: Данные ячейки
    :param data: Formdata
    :return: Описание изображения
    """
    if image_key in cell_data.cells['names']:
        return cell_data.cells['names'].get(image_key)
    material_key, image_id = image_key.split('.')
    if material_key in cell_data.cells['materials']:
        material = data.get(material_key, {})
        if isinstance(material, str):
            material = json.loads(material)
        values = material.get('values', [])
        defect = values[int(image_id)].get('def')
        return defect
    return None


def apply_image_size(user_image, width, height):
    """
    Применяет размеры изображения и конвертирует его в формат JPG.

    :param user_image: Объект UserImage
    :param width: Ширина
    :param height: Высота
    """
    try:
        # Открываем изображение
        img = PILImage.open(user_image.image)
        # print(f"Изображение успешно открыто")
        # Масштабируем изображение
        img = img.resize((width, height), PILImage.LANCZOS)
        # Сохраняем изображение с тем же именем
        img.save(user_image.image, 'JPEG')
        # print(
        #     f"Изображение успешно масштабировано до {width}x{height} и конвертировано в JPG")
    except Exception as e:
        print(
            f"Ошибка при масштабировании и конвертации изображения {user_image.image}: {e}")


def move_image_to_new_line(current_img_cell, current_img_merge_cell,
                           cell_data):
    vertical_gap = int(cell_data.cells['verticalGap'])
    text_height = int(cell_data.cells['textHeight'])
    col_current_img_cell, row_current_img_cell = coordinate_from_string(
        current_img_cell)
    col_current_img_merge_cell, row_current_img_merge_cell = coordinate_from_string(
        current_img_merge_cell)

    # Вычисляем ширину и высоту изображения
    img_width = column_index_from_string(
        col_current_img_merge_cell) - column_index_from_string(
        col_current_img_cell) + 1
    img_height = row_current_img_merge_cell - row_current_img_cell + 1
    # print(f'Image width: {img_width}, Image height: {img_height}')
    # Сдвиг вниз на высоту изображения + 2 + 2 + 1
    new_row = row_current_img_cell + img_height + text_height + vertical_gap

    # Сдвиг влево до столбца указанного в cell_data.index
    new_col = column_index_from_string(
        coordinate_from_string(cell_data.index)[0])
    # Вычисляем новые координаты
    new_img_cell = f"{get_column_letter(new_col)}{new_row}"
    new_img_merge_cell = f"{get_column_letter(new_col + img_width - 1)
    }{new_row + img_height - 1}"

    # print(f'Old cell: {current_img_cell}, New cell: {new_img_cell}')
    # print(
    #     f'Old merge cell: {current_img_merge_cell}, New merge cell: {new_img_merge_cell}')
    return new_img_cell, new_img_merge_cell


def check_image_fits(current_img_cell, current_img_merge_cell, max_cell,
                     cell_data):
    """
    Проверяет, помещается ли изображение на листе.

    :param current_img_cell: Текущая ячейка для вставки изображения (например, "G9")
    :param current_img_merge_cell: Текущая ячейка для вставки изображения (например, "G9")
    :param max_cell: Максимальная ячейка (например, "AM53")
    """
    # Преобразуем max_cell в числовые значения
    max_col, max_row = coordinate_from_string(max_cell)
    max_col_num = column_index_from_string(max_col)

    # Разделяем current_img_merge_cell на буквы и цифры
    col, row = coordinate_from_string(current_img_merge_cell)
    col_num = column_index_from_string(col)

    # Проверяем, превышает ли текущая ячейка максимальные значения
    if row >= max_row:
        return False, current_img_cell, current_img_merge_cell
    elif col_num >= max_col_num:
        new_img_cell, new_img_merge_cell = move_image_to_new_line(
            current_img_cell,
            current_img_merge_cell,
            cell_data)
        return check_image_fits(new_img_cell, new_img_merge_cell, max_cell,
                                cell_data)
    else:
        return True, current_img_cell, current_img_merge_cell


def insert_image_description(ws, user_image, current_img_cell,
                             current_img_merge_cell, cell_data, image_counter):
    """
    Вставляет описание изображения в ячейку под изображением и делает мерж ячеек.

    :param ws: Рабочий лист Excel
    :param user_image: Объект UserImage с изображением и описанием
    :param current_img_cell: Ячейка для вставки изображения (например, "G9")
    :param current_img_merge_cell: Ячейка для вставки изображения (например, "I12")
    """
    desc_merge_range = None
    try:
        text_height = int(cell_data.cells['textHeight'])
        # Получаем координаты начальной и конечной ячеек изображения
        start_col, start_row = coordinate_from_string(current_img_cell)
        end_col, end_row = coordinate_from_string(current_img_merge_cell)

        # Вычисляем высоту и ширину изображения в ячейках
        img_width = column_index_from_string(
            end_col) - column_index_from_string(start_col) + 1

        # Преобразуем start_col в индекс столбца
        desc_start_col = column_index_from_string(start_col)

        # Вычисляем ячейку для описания
        desc_start_row = end_row + 1
        desc_end_row = desc_start_row + 1
        desc_end_col = desc_start_col + img_width - 1

        # Формируем координаты ячейки для описания и диапазона для мержа
        desc_start_cell = f"{get_column_letter(desc_start_col)}{desc_start_row}"
        desc_end_cell = f"{get_column_letter(desc_end_col)}{desc_end_row}"
        desc_merge_range = f"{desc_start_cell}:{desc_end_cell}"

        # Вставляем описание изображения в ячейку
        ws[
            desc_start_cell] = (f"Рисунок Б.{image_counter} –"
                                f" {user_image.description}.")

        # Делаем мерж ячеек для описания
        ws.merge_cells(desc_merge_range)

        # Центрируем текст в ячейке
        ws[desc_start_cell].alignment = Alignment(horizontal='center',
                                                  vertical='center')

        # print(
        #     f"Описание изображения успешно вставлено в диапазон {desc_merge_range}")
    except Exception as e:
        print(
            f"Ошибка при вставке описания изображения {user_image.image} в диапазон {desc_merge_range}: {e}")


def insert_image(ws, user_image, current_img_cell):
    """
    Вставляет изображение в указанную ячейку.

    :param ws: Рабочий лист Excel
    :param user_image: Объект UserImage с изображением и описанием
    :param current_img_cell: Ячейка для вставки изображения (например, "G9")
    """
    try:
        # Открываем изображение
        img = Image(user_image.image)
        # Вставляем изображение в указанную ячейку
        ws.add_image(img, current_img_cell)
        # print(
        #     f"Изображение успешно вставлено в ячейку {current_img_cell}")
    except Exception as e:
        print(
            f"Ошибка при вставке изображения {user_image.image} в ячейку {current_img_cell}: {e}")


def create_new_sheet(ws, ws_initial_copy, sheet_name, copy_count):
    """
    Создает копию изначального листа и изменяет его на текущий рабочий лист.

    :param ws: Рабочий лист Excel
    :param ws_initial_copy: Изначальный лист для копирования
    :param sheet_name: Имя оригинального листа
    :param copy_count: Счетчик копий
    :return: Новый рабочий лист
    """
    new_sheet_name = f"{sheet_name}_{copy_count}"
    new_ws = ws.parent.copy_worksheet(ws_initial_copy)
    new_ws.title = new_sheet_name

    ws.parent._sheets.remove(new_ws)
    previous_index = ws.parent._sheets.index(ws)
    ws.parent._sheets.insert(previous_index + 1, new_ws)

    # print(f"Создан новый лист {new_sheet_name}")
    # print(f'Текущий лист: {new_ws}')
    # print(f'Список листов: {ws.parent.sheetnames}')
    return new_ws, previous_index + 2


def move_image_to_right(current_img_cell, current_img_merge_cell, cell_data):
    horizontal_gap = int(cell_data.cells['horizontalGap'])

    col_current_img_cell, row_current_img_cell = coordinate_from_string(
        current_img_cell)
    col_current_img_merge_cell, row_current_img_merge_cell = coordinate_from_string(
        current_img_merge_cell)

    # Вычисляем ширину изображения
    img_width = column_index_from_string(
        col_current_img_merge_cell) - column_index_from_string(
        col_current_img_cell) + 1

    # Сдвиг вправо на ширину изображения + 1
    new_col = column_index_from_string(
        col_current_img_cell) + img_width + horizontal_gap
    new_row = row_current_img_cell

    # Вычисляем новые координаты
    new_img_cell = f"{get_column_letter(new_col)}{new_row}"
    new_img_merge_cell = f"{get_column_letter(new_col + img_width - 1)}{row_current_img_merge_cell}"

    # print(f'Old cell: {current_img_cell}, New cell: {new_img_cell}')
    # print(
    #     f'Old merge cell: {current_img_merge_cell}, New merge cell: {new_img_merge_cell}')
    return new_img_cell, new_img_merge_cell


def process_images(ws, cell_data, data, sheet,
                   inserted_sheets_count, user_files):
    """
    Заполняет Приложение Б (вставляет файлы в ячейки с подписями).

    :param ws: Рабочий лист Excel <Worksheet "Шаблон">
    :param cell_data: Данные ячейки откуда начинается вставка файлов
    :param data: Formdata
    :param sheet: Лист отчета, объект Sheet
    :param inserted_sheets_count: Количество вставленных листов
    :param user_files: Список файлов пользователя [{"key":"","path":""},...]
    """
    img_width = int(cell_data.cells['width'])
    img_height = int(cell_data.cells['height'])
    image_counter = 1
    # Извлечение данных из sheet.data
    sheet_data = sheet.get_data()
    code_cell = next(
        (item for item in sheet_data if item.inputKey == 'code'), None)
    report_date_cell = next(
        (item for item in sheet_data if item.inputKey == 'reportDate'),
        None)
    code_value = data.get(code_cell.inputKey, '')
    report_date_value = data.get(report_date_cell.inputKey, '')

    user_images = []
    ws_initial_copy = ws.parent.copy_worksheet(ws)
    original_sheet_name = ws.title
    new_sheets_counter = 0
    for user_file in user_files:
        image_key = user_file['key']
        image_path = user_file['path']
        description = get_image_description(image_key, cell_data, data)
        user_image = UserImage(image=image_path, description=description)
        user_images.append(user_image)
    current_img_cell = cell_data.index
    current_img_merge_cell = cell_data.cells['mergeCell']
    max_cell = cell_data.cells['maxCell']
    for user_image in user_images:
        apply_image_size(user_image, img_width, img_height)
        fits, current_img_cell, current_img_merge_cell = check_image_fits(
            current_img_cell,
            current_img_merge_cell,
            max_cell, cell_data)
        if not fits:
            # print(f"Изображение не помещается на листе")
            new_sheets_counter += 1
            ws, idx = create_new_sheet(ws, ws_initial_copy,
                                       original_sheet_name,
                                       new_sheets_counter)

            if sheet.countCell:
                ws[sheet.countCell] = idx
                # print(f"Индекс листа: {idx}")

            # Вставка значений в соответствующие ячейки
            ws[code_cell.index] = code_value
            ws[report_date_cell.index] = report_date_value
            current_img_cell = cell_data.index
            current_img_merge_cell = cell_data.cells['mergeCell']
        # print(f"Изображение помещается на листе")
        insert_image(ws, user_image, current_img_cell)
        ws.merge_cells(f"{current_img_cell}:{current_img_merge_cell}")
        insert_image_description(ws, user_image, current_img_cell,
                                 current_img_merge_cell, cell_data,
                                 image_counter)
        image_counter += 1
        current_img_cell, current_img_merge_cell = move_image_to_right(
            current_img_cell, current_img_merge_cell, cell_data)
    # print(f"Все изображения успешно обработаны")
    return new_sheets_counter + inserted_sheets_count
