import os
import inspect
import pymorphy2
import openpyxl
from datetime import datetime
from apps.excel_app.models import Sheet


def generate_report(data, username):
    template_path = os.path.join(os.path.dirname(__file__), 'report.xlsx')
    wb = openpyxl.load_workbook(template_path)
    count = 0
    for sheet in Sheet.objects.all():
        ws = wb.worksheets[sheet.index]
        count += 1
        for cell_data in sheet.get_data():
            cell = ws[cell_data.index]
            template = cell_data.template
            for key, value in data.items():
                placeholder = f'${key}$'
                if placeholder in template:
                    template = template.replace(placeholder, str(value))

            for part in template.split('$'):
                if '.' in part:
                    word, key = part.split('.')
                    if key in data:
                        initial = data[key]
                        gender = get_gender(initial.split(' ')[-1])
                        if gender:
                            inflected_word = change_gender(word, gender)
                            if inflected_word:
                                template = template.replace(f'${word}.{key}$',
                                                            inflected_word)
            cell.value = template
        try:
            if sheet.countCell:
                ws[sheet.countCell] = count
        except Exception as e:
            print(f'Error updating countCell: {e}')
        sheet.save()

    if wb.worksheets:
        wb.remove(wb.worksheets[-1])
    report_dir = os.path.join(os.path.dirname(__file__), '..', 'media',
                              'reports')
    os.makedirs(report_dir, exist_ok=True)
    report_full_filename = os.path.join(report_dir,
                                        f'{datetime.now().strftime("%H-%M_%d.%m.%Y")}-{username}.xlsx')
    wb.save(report_full_filename)
    # Convert the Excel file to PDF
    os.system(
        f'libreoffice --headless --convert-to pdf {report_full_filename}')
    report_filename = os.path.basename(report_full_filename)
    return report_filename


def patched_getargspec(func):
    fullargspec = inspect.getfullargspec(func)
    return fullargspec.args, fullargspec.varargs, fullargspec.varkw, fullargspec.defaults


inspect.getargspec = patched_getargspec

morph = pymorphy2.MorphAnalyzer()


def get_gender(word):
    parse = morph.parse(word)[0]

    if 'masc' in parse.tag:
        return 'masc'
    elif 'femn' in parse.tag:
        return 'femn'
    elif 'neut' in parse.tag:
        return 'neut'

    return None


def change_gender(word, target_gender):
    parse = morph.parse(word)[0]
    inflected = parse.inflect({target_gender})
    return inflected.word if inflected else None
